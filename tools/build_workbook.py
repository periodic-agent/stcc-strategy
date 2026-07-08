#!/usr/bin/env python3
"""Generate stcc-card-database.xlsx from carddata2.py (v3 layout: Deck column,
extended suits, HYPERLINK image links, status colors, per-tab dropdowns).

WARNING: the live sheet on Periodic_agent's Drive is canonical. Use this only to
bootstrap or to build a merge output; never overwrite volunteer edits
(see WORKFLOW.md, "Card database update cycle").
"""
import sys, os
sys.path.insert(0, os.path.dirname(__file__))
from carddata2 import TBG, SC_ROWS
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

SITE = 'https://periodic-agent.github.io/stcc-strategy/'
OUT = sys.argv[1] if len(sys.argv) > 1 else 'stcc-card-database.xlsx'
wb = Workbook()
F = lambda **k: Font(name='Arial', **k)
HDR = PatternFill('solid', start_color='1F3864')
FILLS = {'AI': PatternFill('solid', start_color='E2EFDA'),
         'im': PatternFill('solid', start_color='FFF2CC'),
         've': PatternFill('solid', start_color='E2EFDA'),
         'ne': PatternFill('solid', start_color='FCE4D6')}
r = wb.active; r.title = 'README'
README = [
 ('ST:CC Community Card Database — To Boldly Go & Second Contact',''),
 ('',''),
 ('Goal','Build the card database for the Card Scanner (stcc-strategy site). One row per card. This sheet only seeds the cards from the strategy guides so far: full market decks, all captain decks and promo packs still need rows.'),
 ('How to help','Pick any row: click its Card image link, compare, fix or fill the fields, put your handle in Contributor (optional), set Status to verified. Add rows for cards we are missing (captain decks especially).'),
 ('Deck',"Common = market cards. Captain deck cards get the captain's name (Georgiou, Soval, Kirk, Archer, Rebner, Khan for Box 2; Pike, Riker, Freeman for Box 3)."),
 ('Suit','Market suits plus Captain, Directive and Status for captain-deck cards (same convention as the Box 1 database).'),
 ('Status colors','Green = pre-filled from card image, needs a human check. Orange = empty, needs full entry. Yellow = image too small to read, needs entry.'),
 ('Traits','Comma-separated, exactly as printed on the card tags. See Vocabulary tab; new expansion traits are welcome, spell as printed.'),
 ('Skill icons','LEFT edge: Research (atom), Influence (handshake), Military (starburst), Any (multicolor), Variable (?). Comma-separated, one per icon.'),
 ('Focus icons','BOTTOM-RIGHT, next to the glory number.'),
 ('Glory','White circle bottom-right. ? = variable; ?* = Rewards card with if-not-logged penalty (see Notes); negatives exist (Incidents).'),
 ('Card code','Bottom-left, e.g. 2PER07/26. OPTIONAL — leave blank if you like. Dagger (†) = updated repeat from the core box.'),
 ('Questions','Ping Periodic_agent.'),
]
for i,(a,b) in enumerate(README,1):
    r[f'A{i}'], r[f'B{i}'] = a, b
    r[f'A{i}'].font = F(bold=True, size=14 if i==1 else 10); r[f'B{i}'].font = F(size=10)
r.column_dimensions['A'].width = 22; r.column_dimensions['B'].width = 130
COLS = ['Card code (optional)','Name','Suit','Deck','Subtype','Species traits','Regular traits','Other traits','Skill icons (left)','Focus icons (bottom-right)','Glory','Card image (click)','Status','Contributor','Notes']
WIDTHS = [16,26,10,11,10,22,32,16,22,20,7,28,34,14,52]

def tab(name, rows, box, captains):
    ws = wb.create_sheet(name)
    for c,(h,w) in enumerate(zip(COLS,WIDTHS),1):
        cell = ws.cell(1,c,h); cell.font = F(bold=True,color='FFFFFF'); cell.fill = HDR
        cell.alignment = Alignment(horizontal='center')
        ws.column_dimensions[get_column_letter(c)].width = w
    ws.freeze_panes = 'A2'
    for i,row in enumerate(rows,2):
        vals = list(row[:3]) + ['Common'] + list(row[3:])
        for c,v in enumerate(vals,1):
            if c == 12 and v:
                cell = ws.cell(i,c,f'=HYPERLINK("{SITE}img/{box}/{v}","{v}")')
                cell.font = F(size=10,color='0563C1',underline='single')
            else:
                ws.cell(i,c,v).font = F(size=10)
        fill = FILLS.get(str(row[11])[:2], FILLS['ne'])
        for c in range(1,16): ws.cell(i,c).fill = fill
    n = len(rows)+120
    for col,vals in [('C','"Person,Ally,Ship,Cargo,Location,Encounter,Incident,Captain,Directive,Status"'),
                     ('D',f'"Common,{captains}"'),
                     ('E','"Starting,Advanced,Rewards"'),
                     ('M','"AI-seeded — verify,verified,needs entry (no image yet),image unreadable — needs manual entry"')]:
        dv = DataValidation(type='list',formula1=vals,allow_blank=True)
        ws.add_data_validation(dv); dv.add(f'{col}2:{col}{n}')

tab('TBG (Box 2)', TBG, 'box2', 'Georgiou,Soval,Kirk,Archer,Rebner,Khan')
tab('Second Contact (Box 3)', SC_ROWS, 'box3', 'Pike,Riker,Freeman')
v = wb.create_sheet('Vocabulary')
SPECIES = "Alien Aenar Andorian Android Bajoran Betazoid Borg Breen Cardassian Changeling Ferengi Human Jem'Hadar Kelpien Klingon Orion Pakled Reman Romulan Synthetic Tellarite Transcendent Trill Vorta Vulcan XB Xindi".split()
REGULAR = ['Ambassador','Anomaly','Augment','Beverage','Business','Cloak','Communication','Creature','Doctor','Dominion','Engineer','Helmet','Hologram','Imperial','Mind Control','Ops','Pilot','Maquis','Scientist','Security','Shady','Spy','Starbase','Starfleet','Telepath','Time Travel','Weapon']
OTHER = ['Attack','Ongoing','Surprise','Wildcard']
NEW = ['NX-01 (TBG)','Ancient (TBG)','Crossover (SC)','Lower Decker (SC)']
ICONS = ['Research','Influence','Military','Any','Variable']
for c,(t,vals) in enumerate([('Species traits (rulebook p.36)',SPECIES),('Regular traits',REGULAR),('Other traits',OTHER),('New expansion traits seen',NEW),('Icon specialties (p.17)',ICONS)],1):
    cell = v.cell(1,c,t); cell.font = F(bold=True,color='FFFFFF'); cell.fill = HDR
    for i,val in enumerate(vals,2): v.cell(i,c,val).font = F(size=10)
    v.column_dimensions[get_column_letter(c)].width = 30
v.freeze_panes = 'A2'
wb.save(OUT)
print('saved', OUT)
