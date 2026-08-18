#!/usr/bin/env python3
"""build_text_from_sheet.py -- import card text from the community sheet into the JSONs.

Round-trip step 2 (see WORKFLOW.md, "Update cycle -- CARD TEXT edition").
Reads the stcc-card-database.xlsx workbook (v2 layout produced by
build_text_workbook.py: one tab per box, Card code in column A, Card text in
the last column) and writes a `strips` array onto every card record in
box1.json .. promo2.json.

The sheet is READ-ONLY; this script never writes to Drive.

Strip schema (agreed with the card-face renderer):
    {"kind": <str>, "action": true|false|null, "qual": <str|null>, "text": <str>}
- kind: lowercase operation keyword (vocabulary below). Hard fail on anything
  else -- same philosophy as build_box2_from_sheet.py's variant markers: the
  one failure that hides itself is a silent fallback.
- action: True if the strip costs an action token (qualifier "Action"),
  False if explicitly "Free", None when the card prints neither.
- qual: verbatim qualifier between the keyword dash and the colon
  ("Action, Requires 3 Military", "Cost", "Bot only", "Attack", ...).
- text: verbatim transcription. Typos ride along until crowd verification.

Rules:
- Ability segments are separated by lines containing only ---.
- Marker segments (bare word: Reserve, Available, ...) duplicate the JSON's
  position_indicator and are dropped (warn on mismatch).
- "THIS CARD CANNOT ..." lines become kind "banner".
- Mission rows (no Card code) have no JSON record yet: skipped, counted.
- Empty text cell -> "strips": [] (untranscribed; Archer/Rebner/Khan decks).
- data/ops.json is (re)written: the kind -> {category, color} map shared with
  the renderer. Colors for the six rulebook strips only; the rest are marked
  "tbd" until checked against scans.

Usage:
  python3 tools/build_text_from_sheet.py --sheet stcc-card-database.xlsx
  (add --dry-run to validate without writing)
"""
import argparse, json, os, re, sys
from openpyxl import load_workbook

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))

TABS = [("Captain's Chair (Box 1)", 'box1.json'),
        ('To Boldly Go (Box 2)',    'box2.json'),
        ('Second Contact (Box 3)',  'box3.json'),
        ('Promo Pack 1',            'promo1.json'),
        ('Promo Pack 2',            'promo2.json')]

# keyword as printed -> kind
KINDS = {
    'Play': 'play', 'Support': 'support', 'Resupply': 'resupply',
    'Clean-Up': 'cleanup', 'Clean-up': 'cleanup', 'Control': 'control',
    'Activation': 'activation', 'Passive': 'passive', 'Reaction': 'reaction',
    'Special': 'special', 'Endgame': 'endgame', 'Development': 'cost',
    'Goal': 'goal', 'Reward': 'reward', 'Away Teams': 'awayteams',
    'Surprise': 'surprise',
}
OPS_JSON = {
    'play':      {'category': 'play',    'color': 'gray'},
    'support':   {'category': 'play',    'color': 'gray'},
    'resupply':  {'category': 'upkeep',  'color': 'green'},
    'cleanup':   {'category': 'upkeep',  'color': 'green'},
    'activation':{'category': 'table',   'color': 'blue'},
    'passive':   {'category': 'table',   'color': 'blue'},
    'reaction':  {'category': 'table',   'color': 'blue'},
    'endgame':   {'category': 'endgame', 'color': 'red'},
    'special':   {'category': 'special', 'color': 'purple'},
    'cost':      {'category': 'devcost', 'color': 'black'},
    'control':   {'category': 'control', 'color': 'tbd'},
    'goal':      {'category': 'mission', 'color': 'tbd'},
    'reward':    {'category': 'mission', 'color': 'tbd'},
    'awayteams': {'category': 'table',   'color': 'tbd'},
    'surprise':  {'category': 'special', 'color': 'tbd'},
    'banner':    {'category': 'banner',  'color': 'none'},
}

MARKERS = {'Reserve', 'Available', 'Development', 'Rewards', 'Status',
           'Captain', 'Deployed', 'Advanced', 'Starting', 'Discard',
           'Incident Deck', 'Controlled Location'}

AWAY_STAT_RE = re.compile(r'^Away Teams:\s*(\d+\+?)$', re.I)

# qualifier grammar: comma-separated parts, each matching one of these
QUAL_PART = re.compile(
    r'^(Action|Free|Cost|Attack|Bot only|Requires .{1,30}|Once per turn)$', re.I)

KEYWORD_RE = re.compile(
    r'^(%s)\s*(?:\(([^)]{1,30})\))?\s*[-–—:]\s*(.*)$' % '|'.join(
        sorted((re.escape(k) for k in KINDS), key=len, reverse=True)),
    re.S | re.I)
KINDS_CI = {k.lower(): v for k, v in KINDS.items()}


def parse_segment(seg, errors, where):
    seg = seg.strip()
    if not seg:
        return None
    if seg in MARKERS:
        return {'_marker': seg}
    st = AWAY_STAT_RE.match(seg)
    if st:
        return {'_stat': ('away_team', st.group(1))}
    if re.match(r'^THIS CARD CANNOT', seg, re.I):
        return {'kind': 'banner', 'action': None, 'qual': None, 'text': seg}
    m = KEYWORD_RE.match(seg)
    if not m:
        errors.append('%s: no operation keyword: %r' % (where, seg[:70]))
        return None
    kind = KINDS_CI[m.group(1).lower()]
    paren_qual = m.group(2)
    rest = m.group(3).strip()
    qual = None
    # a qualifier is the text before the FIRST colon, if every comma-part
    # matches the qualifier grammar
    if ':' in rest:
        cand, after = rest.split(':', 1)
        cand = cand.strip()
        if cand and len(cand) <= 45 and all(
                QUAL_PART.match(p.strip()) for p in cand.split(',')):
            qual, rest = cand, after.strip()
    if paren_qual:
        qual = '%s, %s' % (paren_qual, qual) if qual else paren_qual
    action = None
    if qual:
        parts = [p.strip().lower() for p in qual.split(',')]
        if 'action' in parts:
            action = True
        elif 'free' in parts:
            action = False
    return {'kind': kind, 'action': action, 'qual': qual, 'text': rest}


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--sheet', required=True)
    ap.add_argument('--dry-run', action='store_true')
    args = ap.parse_args()

    wb = load_workbook(args.sheet, read_only=True, data_only=True)
    errors, warnings = [], []
    stats = {'cards': 0, 'with_text': 0, 'empty': 0, 'strips': 0,
             'missions_skipped': 0, 'markers_dropped': 0}

    for tab, jf in TABS:
        cards = json.load(open(os.path.join(ROOT, jf)))
        by_code = {}
        for c in cards:
            by_code.setdefault(c.get('card_number'), []).append(c)
        seen_codes = set()
        ws = wb[tab]
        hdr = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
        text_col = hdr.index('Card text (--- separates abilities)')
        for i, row in enumerate(ws.iter_rows(min_row=2, values_only=True), 2):
            code, name = row[0], row[1]
            if not name:
                continue
            if not code:  # Mission rows: no JSON record yet
                stats['missions_skipped'] += 1
                continue
            pool = by_code.get(code, [])
            # duplicate codes are impossible (codes are unique), but a code
            # seen twice in the sheet is a volunteer paste error
            if code in seen_codes:
                errors.append('%s row %d: duplicate card code %s' % (tab, i, code))
                continue
            seen_codes.add(code)
            if not pool:
                errors.append('%s row %d: card code %r not in %s' % (tab, i, code, jf))
                continue
            card = pool[0] if len(pool) == 1 else next(
                (c for c in pool if c['name'] == name), pool[0])
            stats['cards'] += 1
            cell = (row[text_col] or '').strip()
            strips = []
            if cell:
                for seg in cell.split('---'):
                    s = parse_segment(seg, errors, '%s row %d (%s)' % (tab, i, name))
                    if not s:
                        continue
                    if '_stat' in s:
                        k, v = s['_stat']
                        if k in card and str(card[k]) != v:
                            warnings.append('%s row %d (%s): sheet %s=%r vs '
                                            'JSON %r' % (tab, i, name, k, v, card[k]))
                        else:
                            card[k] = v
                        continue
                    if '_marker' in s:
                        stats['markers_dropped'] += 1
                        pi = card.get('position_indicator')
                        if pi and s['_marker'] not in (pi, pi + 's'):
                            warnings.append('%s row %d (%s): marker %r vs '
                                            'position_indicator %r'
                                            % (tab, i, name, s['_marker'], pi))
                        continue
                    strips.append(s)
            card['strips'] = strips
            if strips:
                stats['with_text'] += 1
                stats['strips'] += len(strips)
            else:
                stats['empty'] += 1
        # cards the sheet never mentioned keep/gain an empty strips field
        for c in cards:
            c.setdefault('strips', [])
        if not args.dry_run and not errors:
            json.dump(cards, open(os.path.join(ROOT, jf), 'w'),
                      indent=1, ensure_ascii=False)

    if not args.dry_run and not errors:
        json.dump(OPS_JSON, open(os.path.join(ROOT, 'data', 'ops.json'), 'w'),
                  indent=1)

    for w in warnings:
        print('WARN  ' + w, file=sys.stderr)
    if errors:
        print('\n%d ERRORS -- nothing written:' % len(errors), file=sys.stderr)
        for e in errors:
            print('  ' + e, file=sys.stderr)
        sys.exit(1)
    print('%(cards)d cards matched | %(with_text)d with text '
          '(%(strips)d strips) | %(empty)d empty | '
          '%(missions_skipped)d mission rows skipped | '
          '%(markers_dropped)d markers dropped' % stats)


if __name__ == '__main__':
    main()
