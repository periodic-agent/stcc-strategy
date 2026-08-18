#!/usr/bin/env python3
"""build_text_workbook.py -- regenerate stcc-card-database.xlsx from the ground up.

Card-TEXT crowdsourcing edition (Aug 2026). Supersedes build_workbook.py.

Reads the five canonical JSON files at repo root (box1.json .. promo2.json)
and the community rule-text transcription (data/card-text-source.csv, exported
from Periodic_agent's Numbers sheet) and emits a workbook with one tab per box,
one row per card, all canonical metadata pre-filled, and **Card text as the
last column** for volunteer verification.

Deck/name matching, not name-only: a sheet text row is joined to a JSON card on
(deck, normalized name). JSON source 'Common'/'Promo' maps to sheet deck
'Market'. Scenario decks (Archer, Rebner, Khan) have no rows in the text
source, so their cards are seeded empty with status 'needs text' -- this also
prevents an Archer 'Analyze' from silently inheriting Soval's text.

Normalization absorbs the source sheet's known spelling drift (Analyse/Analyze,
Lenara Khan/Kahn, ...) via fuzzy fallback at ratio >= 0.90 within the same deck.

Mission cards exist in the text source but not (yet) in the JSONs: they are
appended to each box tab after the JSON cards, marked 'not in JSON yet'.

Usage:
  python3 tools/build_text_workbook.py \
      --text data/card-text-source.csv [-o stcc-card-database.xlsx]

Upload cycle: Periodic_agent replaces the Drive file via right-click ->
Manage versions -> Upload new version (keeps file ID and shared link).
"""
import argparse, csv, json, os, re, sys, unicodedata
from difflib import SequenceMatcher

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.worksheet.datavalidation import DataValidation
from openpyxl.utils import get_column_letter

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
SITE = 'https://periodic-agent.github.io/stcc-strategy/'

BOXES = [  # (json file, tab title, img folder, sheet 'Box' labels used in text source)
    ('box1.json',   "Captain's Chair (Box 1)", 'box1',   {'Core Box'}),
    ('box2.json',   'To Boldly Go (Box 2)',    'box2',   {'To Boldly Go'}),
    ('box3.json',   'Second Contact (Box 3)',  'box3',   {'Second Contact'}),
    ('promo1.json', 'Promo Pack 1',            'promo1', {'Promo 01'}),
    ('promo2.json', 'Promo Pack 2',            'promo2', {'Promo 02'}),
]

STATUS_SEEDED = 'text seeded - please verify vs card'
STATUS_NEEDS = 'needs text (please transcribe)'
STATUS_VERIFIED = 'text verified'
STATUS_NOJSON = 'not in JSON yet (mission)'

COLS = ['Card code', 'Name', 'Suit', 'Deck', 'Variant', 'Position',
        'Species traits', 'Regular traits', 'Other traits',
        'Skill icons (left)', 'Focus icons (bottom-right)', 'Glory',
        'Card image (click)', 'Text status',
        'Card text (--- separates abilities)']
WIDTHS = [13, 26, 11, 10, 9, 10, 18, 24, 14, 18, 18, 7, 26, 30, 95]
TEXT_COL = len(COLS)

FILLS = {
    STATUS_SEEDED:   PatternFill('solid', start_color='E2EFDA'),  # green
    STATUS_NEEDS:    PatternFill('solid', start_color='FCE4D6'),  # orange
    STATUS_VERIFIED: PatternFill('solid', start_color='D9E1F2'),  # blue
    STATUS_NOJSON:   PatternFill('solid', start_color='FFF2CC'),  # yellow
}


def norm(s):
    s = unicodedata.normalize('NFKD', s)
    s = ''.join(c for c in s if not unicodedata.combining(c))
    s = s.replace('’', "'").replace('‘', "'").lower()
    s = re.sub(r'\(.*?\)', '', s)
    s = re.sub(r'\[.*?\]', '', s)
    s = re.sub(r'[^a-z0-9]+', ' ', s).strip()
    return s.replace('analyse', 'analyze').replace('utilise', 'utilize')


def deck_of(source):
    return 'Market' if source in ('Common', 'Promo') else source


def load_text_rows(path):
    with open(path, encoding='utf-8-sig') as f:
        rows = list(csv.reader(f))
    out = []
    for r in rows[1:]:
        if not r or not r[4].strip():
            continue
        boxes = [b.strip() for b in r[1].split('\n') if b.strip()]
        deck = 'Market' if r[2].strip() == 'Market' else r[2].strip()
        out.append({'boxes': boxes, 'deck': deck, 'suit': r[3].strip(),
                    'name': r[4].strip(), 'text': r[6].strip(), 'n': norm(r[4])})
    return out


def match_text(card, deck, box_labels, text_rows, used):
    """Best text row for a JSON card: exact (deck, name), then fuzzy same-deck."""
    cn = norm(card['name'])
    # Mission rows never feed a JSON card (no Mission suit in the JSONs; a deck can
    # hold a Mission and a Directive with the same name, e.g. Koloth's Sabotage)
    rows = [t for t in text_rows if t['suit'] != 'Mission']
    pool = [t for t in rows if t['deck'] == deck and set(t['boxes']) & box_labels]
    # reprints/updates: a Core Box row also serves the Box 2/3 reprint of the same card
    pool_any_box = [t for t in rows if t['deck'] == deck]
    for p in (pool, pool_any_box):
        for t in p:
            if t['n'] == cn:
                used.add(id(t))
                return t
    best, score = None, 0.0
    for t in pool_any_box:
        s = SequenceMatcher(None, cn, t['n']).ratio()
        if s > score:
            best, score = t, s
    if best and score >= 0.85:  # deck-scoped, so a loose ratio cannot cross decks
        used.add(id(best))
        return best
    return None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--text', required=True, help='card text CSV (Numbers export)')
    ap.add_argument('-o', '--out', default='stcc-card-database.xlsx')
    args = ap.parse_args()

    text_rows = load_text_rows(args.text)
    used = set()

    wb = Workbook()
    F = lambda **k: Font(name='Arial', **k)
    HDR = PatternFill('solid', start_color='1F3864')

    # README ---------------------------------------------------------------
    ws = wb.active
    ws.title = 'README'
    README = [
        ('ST:CC Community Card Database v2 - CARD TEXT edition', ''),
        ('', ''),
        ('Goal', 'One row per card, every box. Metadata (name, suit, traits, icons, glory) is '
                 'canonical and pre-filled from the site database; you normally do not need to touch it. '
                 'The job now is the LAST column: Card text. Verify seeded text against the printed card, '
                 'or transcribe it where the row is orange.'),
        ('How to help', 'Pick a row. Click its Card image link, read the card, fix or fill the Card text '
                        'cell, set Text status to "%s".' % STATUS_VERIFIED),
        ('Card text format', 'One ability per line where possible; separate distinct abilities with a line '
                             'containing ---. Copy the card exactly as printed, typos and all: the database '
                             'follows the printed card.'),
        ('Status colors', 'Green = text pre-seeded, needs a human check against the card. '
                          'Orange = no text yet, needs full transcription (Archer / Rebner / Khan decks especially). '
                          'Blue = verified. Yellow = Mission cards not yet in the site database; text captured here first.'),
        ('Deck', "Market = market cards (all boxes). Captain decks carry the captain's name. "
                 'Archer, Rebner and Khan are ordinary decks here, nothing special.'),
        ('Variant', 'original = new in its box; reprint = gameplay-identical repeat of a Box 1 card; '
                    'updated = repeat with new traits or errata (check its text carefully; it may differ from Box 1).'),
        ('Do not', 'Do not sort, insert or delete rows; add corrections in place. Do not edit other tabs.'),
        ('Questions', 'Ping Periodic_agent (BGG).'),
    ]
    for i, (a, b) in enumerate(README, 1):
        ws[f'A{i}'], ws[f'B{i}'] = a, b
        ws[f'A{i}'].font = F(bold=True, size=14 if i == 1 else 10)
        ws[f'B{i}'].font = F(size=10)
        ws[f'B{i}'].alignment = Alignment(wrap_text=True, vertical='top')
    ws.column_dimensions['A'].width = 20
    ws.column_dimensions['B'].width = 130

    counts = {'cards': 0, 'seeded': 0, 'needs': 0, 'missions': 0, 'fuzzy': 0}

    # Box tabs -------------------------------------------------------------
    for jf, tab, imgfolder, box_labels in BOXES:
        cards = json.load(open(os.path.join(ROOT, jf)))
        ws = wb.create_sheet(tab)
        for c, (h, w) in enumerate(zip(COLS, WIDTHS), 1):
            cell = ws.cell(1, c, h)
            cell.font = F(bold=True, color='FFFFFF')
            cell.fill = HDR
            cell.alignment = Alignment(horizontal='center')
            ws.column_dimensions[get_column_letter(c)].width = w
        ws.freeze_panes = 'C2'

        dv = DataValidation(type='list', allow_blank=True,
                            formula1='"%s"' % ','.join([STATUS_SEEDED, STATUS_NEEDS,
                                                        STATUS_VERIFIED, STATUS_NOJSON]))
        ws.add_data_validation(dv)

        r = 1
        for card in cards:
            r += 1
            deck = deck_of(card.get('source', 'Common'))
            t = match_text(card, deck, box_labels, text_rows, used)
            status = STATUS_SEEDED if t else STATUS_NEEDS
            counts['cards'] += 1
            counts['seeded' if t else 'needs'] += 1
            if t and t['n'] != norm(card['name']):
                counts['fuzzy'] += 1
            skill = ', '.join(i['specialty'] for i in card.get('icons', []) if i.get('type') == 'Skill')
            focus = ', '.join(i['specialty'] for i in card.get('icons', []) if i.get('type') == 'Focus')
            vals = [card.get('card_number') or '', card['name'], card['suit'], deck,
                    card.get('variant') or '', card.get('position_indicator') or '',
                    ', '.join(card.get('species_traits', [])),
                    ', '.join(card.get('regular_traits', [])),
                    ', '.join(card.get('other_traits', [])),
                    skill, focus,
                    card['glory'] if card.get('glory') is not None else '',
                    None, status, t['text'] if t else '']
            for c, v in enumerate(vals, 1):
                if c == 13:
                    fn = card.get('filename')
                    cell = ws.cell(r, c, f'=HYPERLINK("{SITE}img/{imgfolder}/{fn}","{fn}")' if fn else '')
                    cell.font = F(size=9, color='0563C1', underline='single')
                else:
                    cell = ws.cell(r, c, v)
                    cell.font = F(size=10)
                cell.alignment = Alignment(wrap_text=(c in (7, 8, TEXT_COL)), vertical='top')
            ws.cell(r, 14).fill = FILLS[status]
            dv.add(ws.cell(r, 14))

        # Mission rows: in the text source but not in the JSONs yet
        for t in text_rows:
            if t['suit'] == 'Mission' and set(t['boxes']) & box_labels:
                r += 1
                counts['missions'] += 1
                used.add(id(t))
                vals = ['', t['name'], 'Mission', t['deck'], '', '', '', '', '', '', '', '',
                        '', STATUS_NOJSON, t['text']]
                for c, v in enumerate(vals, 1):
                    cell = ws.cell(r, c, v)
                    cell.font = F(size=10)
                    cell.alignment = Alignment(wrap_text=(c == TEXT_COL), vertical='top')
                ws.cell(r, 14).fill = FILLS[STATUS_NOJSON]
                dv.add(ws.cell(r, 14))

    # Orphans check --------------------------------------------------------
    orphans = [t for t in text_rows if id(t) not in used and t['suit'] != 'Mission']
    if orphans:
        print('WARNING: %d text rows matched no JSON card:' % len(orphans), file=sys.stderr)
        for t in orphans:
            print('  %-16s %-10s %s' % (t['boxes'][0], t['deck'], t['name']), file=sys.stderr)

    wb.save(args.out)
    print('wrote %s | %d cards (%d text-seeded, %d need text, %d fuzzy joins) '
          '+ %d mission rows | %d orphan text rows'
          % (args.out, counts['cards'], counts['seeded'], counts['needs'],
             counts['fuzzy'], counts['missions'], len(orphans)))


if __name__ == '__main__':
    main()
