#!/usr/bin/env python3
"""
build_box2_from_sheet.py -- build canonical boxN.json from the community sheet.

Step 3 of the "Scanner regeneration checklist" in WORKFLOW.md. Reads the
stcc-card-database.xlsx export (Drive file 186ZpFkLQsLX1blU3z45znMPwH9fE6yO3)
and emits the canonical per-box JSON consumed by tools/build_scanner_data.py.
The sheet is READ-ONLY; this script never writes to Drive.

Sheet columns (TBG (Box 2) / Second Contact (Box 3) tabs):
  A Card code | B Name | C Suit | D Deck | E Subtype | F Species traits
  G Regular traits | H Other traits | I Skill icons | J Focus icons
  K Glory | L Card image | M Status | N Contributor | O Notes

----------------------------------------------------------------------------
CARD NUMBER AND VARIANT
----------------------------------------------------------------------------
Column A carries the printed card number and encodes duplicate status:

    ^(\\d)([A-Z]+)(\\d+)/(\\d+)\\s*(.)?$
       box  set    num  total  marker

    no marker  -> variant "original"
    U+2022 •   -> variant "reprint"   (gameplay-identical to the box 1 card)
    U+2020 †   -> variant "updated"   (new traits or errata)
    anything else in the marker position -> HARD FAIL

The hard fail is deliberate and is not a comment on contributor care. Sheets
autocorrect, mobile keyboards and paste from other sources silently substitute
U+00B7 (·) or U+2219 (∙) for • with no visible change to the typist. Falling back
to "original" is the one failure that hides itself: a † card would look brand new
and the scanner's resolver would never fire on it. All bad markers are collected
and reported together with their codepoints, so one run fixes every cell.

A blank card code is NOT "no marker" -- it is "unclassified". Treating it as
original would reintroduce the same silent failure through a different door.

----------------------------------------------------------------------------
VALIDATION
----------------------------------------------------------------------------
CHECK 1 (gate)  Every • / † row must resolve to an existing box 1 id, joined on
                slug(name). Fatal by default: the whole reason box 2 stores only
                `variant` and no `supersedes` pointer is that this gate guarantees
                the runtime name-join resolves. A stored pointer would bridge a
                name mismatch and let the two spellings drift apart forever
                (this is exactly how "Borg Spatial Trajector[y]" was caught).
                --warn-unresolved downgrades it to a warning to unblock a build.

CHECK 2 (report) The reverse, and the more valuable one: any UNMARKED box 2 card
                whose name matches a box 1 id is a missed transcription, which
                silently double-counts traits in a combined market with nothing
                visibly wrong on screen. Restricted to deck == Common: only market
                cards are reprinted across boxes, while captain decks legitimately
                reuse names (Georgiou's "Hostile Contact", Soval's "Energy Drain"),
                and flagging those every build trains people to ignore the check.

CHECK 3 (report) For every • row, diff the trait fields against the box 1 record.
                They are identical by definition, so any difference means the card
                should have been † -- a change that would otherwise propagate
                silently to box-1-only players. Cheap, so it runs every build.

CHECK 4 (gate)  Identity: a box 2 id collides with a box 1 id IF AND ONLY IF the
                card is marked • or †. Bidirectional, and sharper than 1 and 2
                because it joins on id rather than name. An unmarked collision is a
                missed transcription; a marked card that does NOT collide is naming
                drift between the sheet's image cell and box 1 (this is what caught
                Phlox carrying phlox-nx01.jpg, which checks 1 and 2 both pass over).
                Common-deck rows only, for the reason given under CHECK 2.

Also prints the full • / † list for a human eyeball pass, turning the
classification from an assumption into something verified once.

----------------------------------------------------------------------------
IDS
----------------------------------------------------------------------------
box 1 convention, verified across all 255 core cards: id == filename stem.
Crew-deck cards carry a deck prefix and drop any trailing "(Captain)"
parenthetical; Common/Promo cards take no prefix. When the sheet supplies an
image, that stem is authoritative. `id` is NOT the box-1 join key -- the join is
slug(name) -- so deck-prefixed ids stay unique across captain decks.

Usage:
  python build_box2_from_sheet.py sheet.xlsx "TBG (Box 2)" "To Boldly Go" \
         --box1 box1.json -o box2.json [--warn-unresolved]
"""
import openpyxl, json, re, argparse, sys
from collections import Counter, defaultdict

ICON_SPECIALTIES = {"Research", "Influence", "Military", "Any", "Variable"}
KNOWN_GAME_BOXES = {"Captain's Chair", "To Boldly Go", "Second Contact"}
CODE_RE = re.compile(r"^(\d)([A-Z]+)(\d+)([A-Z]?)/(\d+)\s*(.)?$")  # optional letter = double-sided face (2KHA01A/22)
MARKERS = {None: "original", "•": "reprint", "†": "updated"}



def slug(name):
    s = name.lower()
    for ch in ("'", "’", ".", ","):
        s = s.replace(ch, "")
    return re.sub(r"[^a-z0-9]+", "-", s).strip("-")


def make_id(name, deck, filename):
    if filename:
        return filename[:-4] if filename.lower().endswith(".jpg") else slug(filename)
    base = re.sub(r"\s*\([^)]*\)\s*$", "", name).strip() or name
    if deck and deck not in ("Common", "Promo"):
        return f"{slug(deck)}-{slug(base)}"
    return slug(base)


def splitlist(v):
    return [x.strip() for x in str(v).split(",") if x.strip()] if v is not None else []


def cellstr(ws, r, c):
    v = ws.cell(r, c).value
    return str(v).strip() if v is not None and str(v).strip() else ""


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("xlsx"); ap.add_argument("tab"); ap.add_argument("game_box")
    ap.add_argument("--box1", required=True, help="box1.json, for the duplicate checks")
    ap.add_argument("-o", "--out", required=True)
    ap.add_argument("--warn-unresolved", action="store_true",
                    help="downgrade CHECK 1 from fatal to warning")
    a = ap.parse_args()
    if a.game_box not in KNOWN_GAME_BOXES:
        sys.exit(f"BUILD FAILED -- unknown game_box {a.game_box!r}. "
                 f"Must be one of {sorted(KNOWN_GAME_BOXES)} (mirrors the scanner's GAMEBOX_KEY). "
                 f"An unrecognized value silently resolves to Captain's Chair in the scanner.")

    wb = openpyxl.load_workbook(a.xlsx, data_only=True)
    ws = wb[a.tab]

    voc = wb["Vocabulary"]
    ok = {"species_traits": set(), "regular_traits": set(), "other_traits": set()}
    newexp = set()
    for r in range(2, voc.max_row + 1):
        for col, key in ((1, "species_traits"), (2, "regular_traits"), (3, "other_traits")):
            v = voc.cell(r, col).value
            if v: ok[key].add(str(v).strip())
        v = voc.cell(r, 4).value
        if v: newexp.add(re.sub(r"\s*\(.*?\)", "", str(v)).strip())

    cards, bad_markers, unclassified = [], [], []
    for r in range(2, ws.max_row + 1):
        name, suit = cellstr(ws, r, 2), cellstr(ws, r, 3)
        if not name or not suit:
            continue
        code = cellstr(ws, r, 1)
        variant, marker = "unclassified", None
        if code:
            m = CODE_RE.match(code)
            if not m:
                tail = code[7:] if len(code) > 7 else code
                bad_markers.append((r, name, code,
                                    " ".join(f"U+{ord(c):04X}" for c in tail)))
                continue
            marker = m.group(6)
            if marker not in MARKERS:
                bad_markers.append((r, name, code, f"U+{ord(marker):04X}"))
                continue
            variant = MARKERS[marker]
        else:
            unclassified.append((r, name))

        deck = cellstr(ws, r, 4) or "Common"
        filename = cellstr(ws, r, 12)
        cards.append({
            "id": make_id(name, deck, filename),
            "name": name,
            "suit": suit,
            "source": deck,
            "game_box": a.game_box,
            "card_number": code,
            "variant": variant,
            "species_traits": splitlist(ws.cell(r, 6).value),
            "regular_traits": splitlist(ws.cell(r, 7).value),
            "other_traits": splitlist(ws.cell(r, 8).value),
            "filename": filename,
            "icons": [{"type": "Skill", "specialty": s} for s in splitlist(ws.cell(r, 9).value)]
                   + [{"type": "Focus", "specialty": s} for s in splitlist(ws.cell(r, 10).value)],
        })

    if bad_markers:
        print("BUILD FAILED -- unparseable card number / unknown variant marker:", file=sys.stderr)
        print("  Valid markers: none, U+2022 BULLET, U+2020 DAGGER.", file=sys.stderr)
        for r, name, code, cps in bad_markers:
            print(f"  row {r:<5} {name:<30} {code!r:<20} marker codepoints: {cps}", file=sys.stderr)
        sys.exit(1)

    seen, dup_warn = defaultdict(int), []
    for c in cards:
        seen[c["id"]] += 1
        if seen[c["id"]] > 1:
            new = f'{c["id"]}-{seen[c["id"]]}'
            dup_warn.append(f'{c["id"]!r} ({c["name"]}) -> {new}')
            c["id"] = new

    b1 = json.load(open(a.box1, encoding="utf-8"))
    b1_ids = {c["id"] for c in b1}
    b1_by_slug = {}
    for c in b1:
        b1_by_slug.setdefault(slug(c["name"]), c)

    marked = [c for c in cards if c["variant"] in ("reprint", "updated")]
    unresolved = [c for c in marked if slug(c["name"]) not in b1_ids
                  and slug(c["name"]) not in b1_by_slug]
    missed = [c for c in cards if c["variant"] == "original"
              and c["source"] == "Common" and slug(c["name"]) in b1_ids]

    def traits(c):
        return (sorted(c["species_traits"]), sorted(c["regular_traits"]),
                sorted(c["other_traits"]),
                sorted((i["type"], i["specialty"]) for i in c["icons"]))
    mismarked = []
    for c in marked:
        if c["variant"] != "reprint":
            continue
        ref = b1_by_slug.get(slug(c["name"]))
        if ref and traits(c) != traits(ref):
            mismarked.append((c, ref))

    json.dump(cards, open(a.out, "w", encoding="utf-8"), indent=2, ensure_ascii=False)

    print(f"TOTAL: {len(cards)}")
    print("PER SUIT:", dict(Counter(c["suit"] for c in cards)))
    print("PER DECK:", dict(Counter(c["source"] for c in cards)))
    print("VARIANTS:", dict(Counter(c["variant"] for c in cards)))
    wi = sum(1 for c in cards if c["filename"])
    print(f"IMAGES: {wi} with / {len(cards) - wi} without")
    for w in dup_warn:
        print("WARN duplicate id:", w)
    for r, n in unclassified:
        print(f"WARN unclassified (blank card number): row {r} {n}")

    novel = {k: {t for c in cards for t in c[k] if t not in ok[k]} for k in ok}
    for k, v in novel.items():
        if v:
            print(f"NOVEL {k}: documented={sorted(v & newexp)} UNDOCUMENTED={sorted(v - newexp)}")
    bad_icons = sorted({i["specialty"] for c in cards for i in c["icons"]
                        if i["specialty"] not in ICON_SPECIALTIES})
    print("BAD ICON SPECIALTIES:", bad_icons or "none")

    print(f"\n--- CHECK 3: reprint trait diff vs box 1 ({len(mismarked)} suspect) ---")
    for c, ref in mismarked:
        print(f"  {c['name']}: marked reprint but traits differ from box 1 -> should this be updated?")
        print(f"     box1 {traits(ref)}")
        print(f"     box2 {traits(c)}")

    print(f"\n--- CHECK 2: unmarked Common card matching a box 1 id ({len(missed)} suspect) ---")
    for c in missed:
        print(f"  {c['name']} (id {c['id']}) -- missing a • or † ?")

    print(f"\n--- MARKED CARDS for eyeball pass ({len(marked)}) ---")
    for c in sorted(marked, key=lambda x: (x["variant"], x["name"])):
        print(f"  {c['variant']:9} {c['card_number']:14} {c['name']:30} img={c['filename'] or '(none)'}")

    marked_common = [c for c in cards if c["variant"] in ("reprint", "updated")
                     and c["source"] == "Common"]
    drift = [c for c in marked_common if c["id"] not in b1_ids]
    stray = [c for c in cards if c["variant"] == "original"
             and c["source"] == "Common" and c["id"] in b1_ids]
    print(f"\n--- CHECK 4: marked <-> box1 id collision identity "
          f"({len(drift)} drift, {len(stray)} stray) ---")
    for c in drift:
        print(f"  DRIFT  {c['name']!r} is {c['variant']} but id {c['id']!r} is not a box 1 id"
              f" -- image cell {c['filename'] or '(none)'} likely disagrees with box 1")
    for c in stray:
        print(f"  STRAY  {c['name']!r} is unmarked but id {c['id']!r} collides with box 1")

    print(f"\n--- CHECK 1: marked rows resolving to box 1 ({len(unresolved)} unresolved) ---")
    for c in unresolved:
        print(f"  {c['variant']} {c['name']!r} ({c['card_number']}) has no box 1 counterpart")
    if (drift or stray) and not a.warn_unresolved:
        print("\nBUILD FAILED -- CHECK 4: marked/collision identity broken.\n"
              "Fix the sheet, or re-run with --warn-unresolved to override.", file=sys.stderr)
        sys.exit(3)
    if unresolved and not a.warn_unresolved:
        print("\nBUILD FAILED -- CHECK 1 is a gate: box 2 stores no `supersedes` pointer,\n"
              "so every marked card must resolve by name for the runtime join to be safe.\n"
              "Fix the sheet, or re-run with --warn-unresolved to override.", file=sys.stderr)
        sys.exit(2)

    print(f"\nWrote {a.out}")


if __name__ == "__main__":
    main()
